# backend/report_generator.py
from datetime import datetime, date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from models.internal import (
    User,
    Event,
    Tag,
    EventTagLink,
    Registration,
    Attendance,
    Notification,
)


def generate_excel_report(
    session: Session,
    date_from: date,
    date_to: date,
) -> BytesIO:
    """
    Генерирует Excel-отчёт с данными из всех таблиц за указанный период.

    Args:
        session: Сессия SQLModel для доступа к БД.
        date_from: Начальная дата периода (включительно).
        date_to: Конечная дата периода (включительно).

    Returns:
        BytesIO с содержимым .xlsx файла.
    """
    dt_from = datetime.combine(date_from, datetime.min.time())
    dt_to = datetime.combine(date_to, datetime.max.time())

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    def style_header(ws, headers):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # ------- Лист 1: Пользователи -------
    ws_users = wb.active
    ws_users.title = "Пользователи"
    headers = [
        "ID", "Никнейм", "Имя", "Фамилия", "Отчество",
        "Роль", "Баллы", "Компания", "Создан",
    ]
    style_header(ws_users, headers)

    users = session.exec(
        select(User).where(
            User.created_at >= dt_from, User.created_at <= dt_to
        )
    ).all()
    for row_num, user in enumerate(users, 2):
        ws_users.cell(row=row_num, column=1, value=user.id)
        ws_users.cell(row=row_num, column=2, value=user.nickname)
        ws_users.cell(row=row_num, column=3, value=user.firstname)
        ws_users.cell(row=row_num, column=4, value=user.lastname)
        ws_users.cell(row=row_num, column=5, value=user.middlename)
        ws_users.cell(row=row_num, column=6, value=user.role.value)
        ws_users.cell(row=row_num, column=7, value=user.points)
        ws_users.cell(row=row_num, column=8, value=user.company)
        ws_users.cell(
            row=row_num, column=9,
            value=user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "",
        )
    auto_width(ws_users)

    # ------- Лист 2: События -------
    ws_events = wb.create_sheet("События")
    headers = [
        "ID", "Название", "Описание", "Дата", "Баллы",
        "Ссылка", "Архив", "Создано",
    ]
    style_header(ws_events, headers)

    events = session.exec(
        select(Event).where(
            Event.created_at >= dt_from, Event.created_at <= dt_to
        )
    ).all()
    for row_num, event in enumerate(events, 2):
        ws_events.cell(row=row_num, column=1, value=event.id)
        ws_events.cell(row=row_num, column=2, value=event.title)
        ws_events.cell(row=row_num, column=3, value=event.description)
        ws_events.cell(
            row=row_num, column=4,
            value=event.date.strftime("%Y-%m-%d %H:%M") if event.date else "",
        )
        ws_events.cell(row=row_num, column=5, value=event.points)
        ws_events.cell(row=row_num, column=6, value=event.link)
        ws_events.cell(
            row=row_num, column=7,
            value="Да" if event.is_archived else "Нет",
        )
        ws_events.cell(
            row=row_num, column=8,
            value=event.created_at.strftime("%Y-%m-%d %H:%M") if event.created_at else "",
        )
    auto_width(ws_events)

    # ------- Лист 3: Теги и связи -------
    ws_tags = wb.create_sheet("Теги")
    headers = [
        "ID тега", "Название тега", "ID события", "Название события",
    ]
    style_header(ws_tags, headers)

    tag_links = session.exec(
        select(EventTagLink, Tag, Event)
        .join(Tag, EventTagLink.tag_id == Tag.id)
        .join(Event, EventTagLink.event_id == Event.id)
        .where(
            EventTagLink.event_id.in_(
                select(Event.id).where(
                    Event.created_at >= dt_from, Event.created_at <= dt_to
                )
            )
        )
    ).all()
    for row_num, (link, tag, event) in enumerate(tag_links, 2):
        ws_tags.cell(row=row_num, column=1, value=tag.id)
        ws_tags.cell(row=row_num, column=2, value=tag.title)
        ws_tags.cell(row=row_num, column=3, value=event.id)
        ws_tags.cell(row=row_num, column=4, value=event.title)
    auto_width(ws_tags)

    # ------- Лист 4: Регистрации -------
    ws_regs = wb.create_sheet("Регистрации")
    headers = [
        "ID пользователя", "Никнейм", "ID события",
        "Название события", "Дата регистрации",
    ]
    style_header(ws_regs, headers)

    registrations = session.exec(
        select(Registration, User, Event)
        .join(User, Registration.user_id == User.id)
        .join(Event, Registration.event_id == Event.id)
        .where(
            Registration.created_at >= dt_from,
            Registration.created_at <= dt_to,
        )
    ).all()
    for row_num, (reg, user, event) in enumerate(registrations, 2):
        ws_regs.cell(row=row_num, column=1, value=user.id)
        ws_regs.cell(row=row_num, column=2, value=user.nickname)
        ws_regs.cell(row=row_num, column=3, value=event.id)
        ws_regs.cell(row=row_num, column=4, value=event.title)
        ws_regs.cell(
            row=row_num, column=5,
            value=reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else "",
        )
    auto_width(ws_regs)

    # ------- Лист 5: Посетители -------
    ws_att = wb.create_sheet("Посетители")
    headers = [
        "ID пользователя", "Никнейм", "ID события",
        "Название события", "Дата отметки",
    ]
    style_header(ws_att, headers)

    attendances = session.exec(
        select(Attendance, User, Event)
        .join(User, Attendance.user_id == User.id)
        .join(Event, Attendance.event_id == Event.id)
        .where(
            Attendance.created_at >= dt_from,
            Attendance.created_at <= dt_to,
        )
    ).all()
    for row_num, (att, user, event) in enumerate(attendances, 2):
        ws_att.cell(row=row_num, column=1, value=user.id)
        ws_att.cell(row=row_num, column=2, value=user.nickname)
        ws_att.cell(row=row_num, column=3, value=event.id)
        ws_att.cell(row=row_num, column=4, value=event.title)
        ws_att.cell(
            row=row_num, column=5,
            value=att.created_at.strftime("%Y-%m-%d %H:%M") if att.created_at else "",
        )
    auto_width(ws_att)

    # ------- Лист 6: Уведомления -------
    ws_notif = wb.create_sheet("Уведомления")
    headers = ["ID", "Заголовок", "Текст", "Создано"]
    style_header(ws_notif, headers)

    notifications = session.exec(
        select(Notification).where(
            Notification.created_at >= dt_from,
            Notification.created_at <= dt_to,
        )
    ).all()
    for row_num, notif in enumerate(notifications, 2):
        ws_notif.cell(row=row_num, column=1, value=notif.id)
        ws_notif.cell(row=row_num, column=2, value=notif.title)
        ws_notif.cell(row=row_num, column=3, value=notif.body)
        ws_notif.cell(
            row=row_num, column=4,
            value=notif.created_at.strftime("%Y-%m-%d %H:%M") if notif.created_at else "",
        )
    auto_width(ws_notif)

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    wb.save("/tmp/test_report.xlsx")
    output.seek(0)
    return output